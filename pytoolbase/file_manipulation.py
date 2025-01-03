import logging
import csv
from datetime import datetime, timedelta
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import os
from shutil import copyfile
from .my_logger import CustomLogger
import formulas
import sys
import glob


class CustomFile:
    """Class to create and style custom files from SQL queries.

    Attributes:
        custom_logger   Instance of the custom logger class. Used for logging purposes
        db_class        Instance of the Database class. Used to get data from the database
    """
    __custom_logger = None
    __db_class = None
    __cfg_class = None

    def __init__(self, configuration_class):
        """At class initialization make sure that the working folder for temporary files
        exists. If not, it creates it
        """
        self.__cfg_class = configuration_class
        self.__custom_logger = CustomLogger('CustomFileClass').custom_logger(
            self.__cfg_class.get_log_level_from_log_key("custom_file_log_level")
        )
        self.__custom_logger.info(f'Initializing CustomFile Class')

        if not os.path.exists(r'.\working_files'):
            os.mkdir(r'.\working_files')

    def create_csv_file_from_pandas_dataframe(self, pandas_dataframe, csv_file_path):
        """Creates a csv file from a specific sql query.


        Args:
            pandas_dataframe (dataframe):   Pandas dataframe to be turned into a csv file
            csv_file_path (str):    Path of the csv file that has to be created

        Return:
            Nothing
        """
        self.__custom_logger.info(f"create_csv_file_from_pandas_dataframe. Parameters: {csv_file_path}")
        self.__custom_logger.debug(f"Dataframe {pandas_dataframe}.")
        
        pandas_dataframe.to_csv(
            csv_file_path,
            encoding='utf-8',
            header=True,
            doublequote=True,
            sep=',',
            index=False,
            decimal='.'
        )
        self.__custom_logger.debug(f"Created csv file {csv_file_path}.")

    def create_excel_file_from_csv(self, csv_file_path, excel_file_path, encoding):
        self.__custom_logger.info(
            f"create_excel_file_from_csv. Parameters: {csv_file_path}, {excel_file_path}, {encoding}"
        )

        wb = Workbook()
        ws = wb.active

        with open(csv_file_path, encoding=encoding) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

        logging.debug("Style the file")
        ws.freeze_panes = ws['A2']
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows():

            for cell in row:

                ws.column_dimensions[cell.column_letter].width = 25

                if cell.row == 1:
                    ws[cell.coordinate].font = Font(bold=True)

                # Cell formatting
                match cell.column_letter:

                    case _:
                        if cell.row > 1:
                            ws[cell.coordinate].number_format = '@'

        self.__custom_logger.info("Deleting the csv file")
        os.remove(csv_file_path)

        self.__custom_logger.info(f"Saving the excel file at {excel_file_path}")
        wb.save(excel_file_path)

        self.__custom_logger.info(f"Excel file created.")

    def copy_file_to_remote_folder(self, local_path, remote_path):
        """Use the copyfile utility to upload a local file to a remote folder

        Args:
            local_path  Path to the local file to be uploaded
            remote_path Path where to copy the local file
        """
        self.__custom_logger.info(f"copy_file_to_remote_folder. Parameters: {local_path}, {remote_path}")
        
        copyfile(local_path, remote_path)

    def get_column_values_for_each_row(self, excel_file_path, column_indexes):
        """

        Args:
            column_indexes (list):  List of column indexes to consider in the selection
            excel_file_path (str):  Path to the Excel file to open

        Returns:
            values_dict (dict): A dictionary containing relevant values
                                retrieved from specific columns
        """
        self.__custom_logger.info(f"get_column_values_for_each_row. Parameters: {excel_file_path}, {column_indexes}")

        workbook = load_workbook(excel_file_path)
        sheet = workbook['Sheet']
        values_dict = {}
        i = 0

        for row in sheet.iter_rows():
            value_list = []
            for cell in row:
                if cell.row != 1:
                    if cell.column in column_indexes:
                        value_list.append(cell.value.replace(' ', ''))

            if len(value_list) != 0:
                values_dict[i] = value_list

            i += 1

        return values_dict

    def save_calculated_cell_value(self, excel_file_to_update):
        """
        Given a specific cell value calculated with a formula,
        saves the value itself without the calculations.

        During the process the initial formula is lost
        """
        self.__custom_logger.info(
            f"save_calculated_cell_value. Parameters: {excel_file_to_update}"
        )

        # Save calculated results
        self.__custom_logger.info(f'Saving calculated results')
        xl_model = formulas.ExcelModel().loads(excel_file_to_update).finish()
        xl_model.calculate()
        xl_model.write(dirpath=r'.\external_files')

    def get_list_from_env_file(self, environment_file_path, key_to_extract):
        self.__custom_logger.info(
            f"get_list_from_env_file. Parameters: {environment_file_path}, {key_to_extract}"
        )

        env_file = self.__cfg_class.load_env_file(environment_file_path)
        values = env_file[key_to_extract]
        list_of_values = list(values)

        for val in list_of_values:
            if val == ',':
                list_of_values.remove(val)

        return list_of_values


class FilesAndFoldersManipulations:
    __custom_file_class = None
    __config_class = None
    __custom_logger = None
    __check_folders = False

    def __init__(self, configuration_class, custom_file):
        self.__custom_file_class = custom_file
        self.__config_class = configuration_class
        self.__custom_logger = CustomLogger('FilesAndFoldersManipulations').custom_logger(
            self.__config_class.get_log_level_from_log_key("files_and_folders_log_level")
        )
        self.__custom_logger.info(f'Initializing FilesAndFoldersManipulations Class')

    def mount_disk(self, network_drive_letter, network_path, username, password):
        self.__custom_logger.info(f'mount_disk')

        os.system(f"net use {network_drive_letter}: {network_path} /user:{username} {password} /persistent:no")

        self.__custom_logger.info(f'Checking if {network_drive_letter} disk has been mounted.')

        network_pt = f'{network_drive_letter}'.strip()

        if os.path.exists(network_pt):
            self.__custom_logger.critical(f'ERROR. Disk {network_drive_letter} could not be mounted. Ending program')
            sys.exit(99)

        self.__custom_logger.info(f'{network_drive_letter} disk is present and ready to be used.')

    def unmount_disk(self, network_drive_letter):
        self.__custom_logger.info(f'unmount_disk')
        os.system(f"net use {network_drive_letter}: /delete /yes")

        self.__custom_logger.info(f'Checking if {network_drive_letter} disk has been unmounted.')

        if not os.path.exists(network_drive_letter):
            self.__custom_logger.info(f'Disk {network_drive_letter} has been removed')

    def ending_program(self):
        self.__custom_logger.info(f'All operations completed. Ending program.')

    def check_folders(self, general_env_file_path):
        self.__custom_logger.info(f'check_folders. Parameters: {general_env_file_path}')

        general_env_secrets = self.__config_class.get_value_from_env_file(general_env_file_path)
        backup_folder = general_env_secrets["backup_folder"]

        self.__custom_logger.info(f'Checking if {backup_folder} exists.')
        if not os.path.exists(backup_folder):
            os.mkdir(backup_folder)

        self.__check_folders = True

    def copy_files_in_folders(self, local_excel_path, remote_file_path, general_env_file_path):
        self.__custom_logger.info(
            f'copy_files_in_folders. Parameters: {local_excel_path}, {remote_file_path}, {general_env_file_path}'
        )

        if not self.__check_folders:
            self.check_folders(general_env_file_path)

        self.__custom_file_class.copy_file_to_remote_folder(local_excel_path, remote_file_path)

    def delete_logs(self):
        self.__custom_logger.info(f'delete_logs')
        yesterday = (datetime.now() + timedelta(days=-1)).strftime("%Y-%m-%d")

        for fl in glob.glob(f'./logs/{yesterday}-*'):
            self.__custom_logger.info(f'Deleting file {fl}')
            os.remove(fl)

    def delete_working_files(self):
        self.__custom_logger.info(f'delete_working_files')

        for fl in glob.glob('./external_files/*.xlsx'):
            self.__custom_logger.info(f'Deleting file {fl}')
            os.remove(fl)
