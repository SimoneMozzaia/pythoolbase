import logging
import csv
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import os
from .database_connection import Queries
from shutil import copyfile


class CustomFile:
    """Class to create and style custom files from SQL queries.

    Attributes:
        custom_logger   Instance of the custom logger class. Used for logging purposes
        db_class        Instance of the Database class. Used to get data from the database
    """
    __custom_logger = None
    __db_class = None

    def __init__(self):
        self.__custom_logger = logging.getLogger(__name__)
        self.__db_class = Queries()

    def create_csv_file_from_sql_query(self, connection, query_path, csv_file_path):
        """Creates a csv file from a specific sql query.


        Args:
            connection (connection): Connection to the database
            query_path (str):   Path to the file in which the queries to call are stored
            csv_file_path (str):    Path of the csv file that has to be created

        Return:
            Nothing
        """
        self.__custom_logger.info("Create csv file")
    
        df = self.__db_class.get_pandas_df_from_query(query_path, connection)
        self.__custom_logger.debug(f"Dataframe {df}.")
        
        df.to_csv(
            csv_file_path
            , encoding='utf-8'
            , header=True
            , doublequote=True
            , sep=','
            , index=False
            , decimal='.'
        )
        self.__custom_logger.debug(f"Created csv file {csv_file_path}.")

        connection.close()
        self.__custom_logger.debug(f"Closed database connection {connection}.")

    def create_excel_file_from_csv(self, csv_file_path, excel_file_path):
        self.__custom_logger.info("Create excel file")

        wb = Workbook()
        ws = wb.active

        with open(csv_file_path) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

        logging.debug("Style the file")
        ws.freeze_panes = ws['A2']
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows():

            for cell in row:

                ws.column_dimensions[cell.column_letter].width = 25

                if cell.row == 1:
                    ws[cell.coordinate].font = Font(bold=True)

                # Cell formatting
                match cell.column_letter:

                    case _:
                        if cell.row > 1:
                            ws[cell.coordinate].number_format = '@'

        self.__custom_logger.debug("Delete the csv file")
        os.remove(csv_file_path)

        self.__custom_logger.debug(f"Saving excel file at {excel_file_path}")
        wb.save(excel_file_path)

        self.__custom_logger.info(f"Operations completed. Closing app.")

    def copy_file_to_remote_folder(self, local_path, remote_path):
        """Use the copyfile utility to upload a local file to a remote folder

        Args:
            local_path  Path to the local file to be uploaded
            remote_path Path where to copy the local file
        """
        self.__custom_logger.info(f"copy_file_to_remote_folder")
        
        copyfile(local_path, remote_path)
