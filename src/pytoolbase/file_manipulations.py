from __future__ import annotations

import csv
from pathlib import Path
from shutil import copy2
from typing import Any, Iterable

from ._optional import OptionalDependencyError, require_dependency
from .configuration_file import EnvFile


def copy_file(src: Path, dst: Path, *, overwrite: bool = True) -> Path:
    """Copy a file preserving metadata."""
    src = src.expanduser().resolve()
    dst = dst.expanduser().resolve()

    if not src.exists():
        raise FileNotFoundError(f"Source file not found: {src}")
    if dst.exists() and not overwrite:
        raise FileExistsError(f"Destination already exists: {dst}")

    copy2(src, dst)
    return dst


def dataframe_to_csv(df: Any, csv_path: Path, *, encoding: str = "utf-8", index: bool = False) -> Path:
    """Save a pandas-like dataframe to CSV.

    This function is intentionally dependency-agnostic: any object implementing
    `.to_csv(...)` is accepted.
    """
    if not hasattr(df, "to_csv"):
        raise TypeError("df must provide a to_csv(...) method")
    csv_path = csv_path.expanduser().resolve()
    df.to_csv(csv_path, encoding=encoding, index=index)
    return csv_path


def csv_to_xlsx(
    csv_path: Path,
    xlsx_path: Path,
    *,
    encoding: str = "utf-8",
    bold_header: bool = True,
) -> Path:
    """Convert a CSV file to a simple XLSX workbook."""
    openpyxl = require_dependency("openpyxl", extra="excel", feature="CSV->XLSX conversion")
    from openpyxl.styles import Font  # type: ignore
    from openpyxl.workbook import Workbook  # type: ignore

    csv_path = csv_path.expanduser().resolve()
    xlsx_path = xlsx_path.expanduser().resolve()

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")

    wb = Workbook()
    ws = wb.active

    with csv_path.open("r", encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            ws.append(row)
            if row_idx == 1 and bold_header:
                for cell in ws[row_idx]:
                    cell.font = Font(bold=True)

    wb.save(xlsx_path)
    return xlsx_path


def read_excel_columns(excel_path: Path, column_indexes: Iterable[int]) -> list[list[Any]]:
    """Return a list of rows, each row containing values of the requested columns.

    `column_indexes` are 1-based indexes (as in Excel).
    """
    require_dependency("openpyxl", extra="excel", feature="Excel reading")
    from openpyxl import load_workbook  # type: ignore

    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    indexes = list(column_indexes)
    if not indexes or any(i <= 0 for i in indexes):
        raise ValueError("column_indexes must contain positive (1-based) integers")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        extracted = []
        for idx in indexes:
            extracted.append(row[idx - 1] if idx - 1 < len(row) else None)
        rows.append(extracted)
    return rows


def save_calculated_cells(excel_path: Path) -> Path:
    """Force formula calculation and save the workbook.

    Requires `xlwings` (Excel must be available on the host).
    """
    try:
        xw = require_dependency("xlwings", extra="excel", feature="Excel formula calculation")
    except OptionalDependencyError as exc:
        raise exc

    excel_path = excel_path.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    app = xw.App(visible=False)
    try:
        book = app.books.open(str(excel_path))
        try:
            app.calculate()
            book.save()
        finally:
            book.close()
    finally:
        app.quit()
    return excel_path


def parse_env_csv_list(env: EnvFile | Path, key: str) -> list[str]:
    """Parse a comma-separated list from a dotenv key."""
    env_file = env if isinstance(env, EnvFile) else EnvFile(env)
    raw = env_file.require(key)
    parts = [p.strip() for p in raw.split(",")]
    return [p for p in parts if p]
